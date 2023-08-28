# pipenv install openpyxl

import openpyxl

# wb = openpyxl.Workbook()

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

# get content of cell:
cell = sheet["a1"]
cell.value
cell.value = 'xxx'

cell.row
cell.column
cell.coordinate

# or:
cell = sheet.cell(row=1, column=1)

for row in range(1, sheet.max_row + 1):
  for col in range(1, sheet.max_column + 1):
    cell = sheet.cell(row, col)
    print(cell.value)

# all cells in "A" column:
sheet["a"]

sheet["a:c"]

sheet["a1: c3"]

# all cells in the 1st row:
sheet[1]
# first 3 rows:
sheet[1:3]

sheet.append([1,2,3])
sheet.insert_rows(...)

wb.save("transactions2.xlsx")


